import openpyxl

workbook = openpyxl.Workbook()

sheet = workbook.active

sheet.title = '蔡徐坤篮球'

sheet.cell(row=1, column=1, value='名称')
sheet.cell(row=1, column=2, value='地址')
sheet.cell(row=1, column=3, value='描述')
sheet.cell(row=1, column=4, value='观看次数')
sheet.cell(row=1, column=5, value='弹幕数')
sheet.cell(row=1, column=6, value='发布时间')

workbook.save('蔡徐坤篮球.xlsx')
