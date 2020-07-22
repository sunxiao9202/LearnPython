import multiprocessing
import time
import requests
from bs4 import BeautifulSoup
import openpyxl

from com.learn.python.utils import IpProxy


def request_douban(url):
    try:
        proxies = IpProxy.get_proxy()
        response = requests.get(url, headers={
            'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0)',
        }, proxies=proxies)
        if response.status_code == 200:
            return response.text
    except requests.RequestException:
        return None


book = openpyxl.Workbook()
sheet = book.active
sheet.title = '豆瓣电影Top250'
sheet.cell(row=1, column=1, value='名称')
sheet.cell(row=1, column=2, value='图片')
sheet.cell(row=1, column=3, value='排名')
sheet.cell(row=1, column=4, value='评分')
sheet.cell(row=1, column=5, value='作者')
sheet.cell(row=1, column=6, value='简介')

n = 2


def save_to_excel(soup):
    list = soup.find(class_='grid_view').find_all('li')

    for item in list:
        item_name = item.find(class_='title').string
        item_img = item.find('a').find('img').get('src')
        item_index = item.find(class_='').string
        item_score = item.find(class_='rating_num').string
        item_author = item.find('p').text
        if (item.find(class_='inq') != None):
            item_intr = item.find(class_='inq').string

        print('爬取电影：' + item_index + ' | ' + item_name + ' | ' + item_score + ' | ' + item_intr)

        global n

        sheet.cell(row=n, column=1, value=item_name)
        sheet.cell(row=n, column=2, value=item_img)
        sheet.cell(row=n, column=3, value=item_index)
        sheet.cell(row=n, column=4, value=item_score)
        sheet.cell(row=n, column=5, value=item_author)
        sheet.cell(row=n, column=6, value=item_intr)

        n = n + 1


def main(url):
    html = request_douban(url)
    if (html):
        soup = BeautifulSoup(html, 'lxml')
        save_to_excel(soup)


if __name__ == '__main__':
    start = time.time()
    urls = []
    pool = multiprocessing.Pool(multiprocessing.cpu_count())
    for i in range(0, 10):
        url = 'https://movie.douban.com/top250?start=' + str(i * 25) + '&filter='
        urls.append(url)
    pool.map(main, urls)
    pool.close()
    pool.join()
    end = time.time()
    print("耗时：" + str(end - start))

book.save(u'豆瓣最受欢迎的250部电影.xlsx')
