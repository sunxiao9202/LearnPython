import time

import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# browser = webdriver.Chrome()
browser = webdriver.PhantomJS()
WAIT = WebDriverWait(browser, 10)
browser.set_window_size(1400, 900)

workbook = openpyxl.Workbook()

sheet = workbook.active
sheet.title = '蔡徐坤篮球'
sheet.cell(row=1, column=1, value='名称')
sheet.cell(row=1, column=2, value='地址')
sheet.cell(row=1, column=3, value='描述')
sheet.cell(row=1, column=4, value='观看次数')
sheet.cell(row=1, column=5, value='弹幕数')
sheet.cell(row=1, column=6, value='发布时间')
n = 2


def search():
    try:
        print("开始访问网站")
        browser.get("https://www.bilibili.com/")
        input = WAIT.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#nav_searchform > input")))
        submit = WAIT.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#nav_searchform .nav-search-btn button')))
        input.send_keys('蔡徐坤 篮球')
        submit.click()

        # 跳转新页面
        print("跳转新页面")
        all_h = browser.window_handles
        browser.switch_to.window(all_h[1])
        get_source()

        total = WAIT.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".last button")))
        return int(total.text)
    except TimeoutException:
        return search()


def next_page(page_num):
    try:
        print("获取下一页数据：" + str(page_num))
        next_btn = WAIT.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='nav-btn iconfont icon-arrowdown3']")))
        next_btn.click()
        WAIT.until(
            EC.text_to_be_present_in_element((By.CSS_SELECTOR, "[class='page-item active'] > button"), str(page_num)))
        get_source()
    except TimeoutException:
        browser.refresh()
        return next_page(page_num)


def get_source():
    WAIT.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#all-list > div.flow-loader > div.filter-wrap')))
    html = browser.page_source
    soup = BeautifulSoup(html, 'lxml')
    save_to_excel(soup)


def save_to_excel(soup):
    list = soup.find(class_='video-list clearfix').find_all(class_='video-item matrix')
    for item in list:
        item_title = item.find('a').get('title')
        item_link = item.find('a').get('href')
        item_dec = item.find(class_='des hide').text
        item_view = item.find(class_='so-icon watch-num').text
        item_biubiu = item.find(class_='so-icon hide').text
        item_date = item.find(class_='so-icon time').text
        print('爬取:' + item_title)

        global n
        sheet.cell(row=n, column=1, value=item_title)
        sheet.cell(row=n, column=2, value=item_link)
        sheet.cell(row=n, column=3, value=item_dec)
        sheet.cell(row=n, column=4, value=item_view)
        sheet.cell(row=n, column=5, value=item_biubiu)
        sheet.cell(row=n, column=6, value=item_date)
        n = n + 1


def main():
    try:
        total = search()
        for i in range(2, total + 1):
            time.sleep(2)
            next_page(i)
    finally:
        pass


if __name__ == '__main__':
    main()
    workbook.save('蔡徐坤篮球.xlsx')
